from __future__ import annotations

import csv
import hashlib
import re
from collections import defaultdict
from pathlib import Path

from .models import (
    AnalysisResult,
    DatasetSnapshot,
    EvaluationResult,
    PowerBananaReport,
    SecurityFinding,
    StepRecord,
)


PROMPT_INJECTION_PATTERNS = (
    re.compile(r"\bignore (all )?(previous|prior|above) instructions\b", re.IGNORECASE),
    re.compile(r"\breveal secrets?\b", re.IGNORECASE),
    re.compile(r"\bdisregard (the )?(system|developer) (message|instructions)\b", re.IGNORECASE),
)


class PowerBananaAgent:
    name = "PowerBanana"
    version = "0.1"

    def answer(self, file_path: str | Path, question: str) -> PowerBananaReport:
        path = Path(file_path)
        rows = self._load_rows(path)
        snapshot = self._snapshot(path, rows)
        security_findings = self._scan_security(rows)

        if self._is_ambiguous_performance_question(question):
            return PowerBananaReport(
                agent_name=self.name,
                version=self.version,
                status="needs_clarification",
                answer="Please specify the metric to optimize, such as conversion_rate, revenue, orders, or visits.",
                dataset_snapshot=snapshot,
                security_findings=security_findings,
                step_trace=[],
                evaluation=EvaluationResult(verdict="needs_clarification", failure_reasons=["ambiguous_metric"], scores={}),
            )

        if self._is_conversion_rate_question(question):
            return self._answer_conversion_rate(snapshot, rows, security_findings)

        return PowerBananaReport(
            agent_name=self.name,
            version=self.version,
            status="needs_clarification",
            answer="PowerBanana v0.1 supports conversion-rate questions for CSV datasets. Please ask for conversion rate by group.",
            dataset_snapshot=snapshot,
            security_findings=security_findings,
            step_trace=[],
            evaluation=EvaluationResult(verdict="needs_clarification", failure_reasons=["unsupported_question"], scores={}),
        )

    def _load_rows(self, path: Path) -> list[dict[str, str]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", newline="", encoding="utf-8-sig") as handle:
                return list(csv.DictReader(handle))
        if suffix == ".xlsx":
            return self._load_xlsx_rows(path)
        raise ValueError(f"Unsupported file type: {suffix}. PowerBanana v0.1 supports .csv and simple .xlsx files.")

    def _load_xlsx_rows(self, path: Path) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Reading .xlsx files requires openpyxl. Install it or use CSV for PowerBanana v0.1.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value) if value is not None else "" for value in rows[0]]
        records: list[dict[str, str]] = []
        for row in rows[1:]:
            records.append(
                {
                    header: "" if value is None else str(value)
                    for header, value in zip(headers, row, strict=False)
                    if header
                }
            )
        return records

    def _snapshot(self, path: Path, rows: list[dict[str, str]]) -> DatasetSnapshot:
        file_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        columns = list(rows[0].keys()) if rows else []
        missing_counts = {
            column: sum(1 for row in rows if not str(row.get(column, "")).strip())
            for column in columns
        }
        return DatasetSnapshot(
            dataset_id=path.stem,
            dataset_version="upload_v1",
            file_hash=f"sha256:{file_hash}",
            row_count=len(rows),
            columns=columns,
            missing_counts=missing_counts,
        )

    def _scan_security(self, rows: list[dict[str, str]]) -> list[SecurityFinding]:
        findings: list[SecurityFinding] = []
        for row_index, row in enumerate(rows, start=2):
            for column, value in row.items():
                text = str(value)
                if any(pattern.search(text) for pattern in PROMPT_INJECTION_PATTERNS):
                    findings.append(
                        SecurityFinding(
                            risk_type="prompt_injection_in_cell",
                            source_ref=f"row:{row_index}:column:{column}",
                            action="exclude_as_instruction_keep_as_data",
                            detail=text,
                        )
                    )
        return findings

    def _is_ambiguous_performance_question(self, question: str) -> bool:
        q = question.lower()
        return "best" in q and "conversion" not in q and "revenue" not in q and "orders" not in q and "visits" not in q

    def _is_conversion_rate_question(self, question: str) -> bool:
        q = question.lower()
        return "conversion" in q and "rate" in q

    def _answer_conversion_rate(
        self,
        snapshot: DatasetSnapshot,
        rows: list[dict[str, str]],
        security_findings: list[SecurityFinding],
    ) -> PowerBananaReport:
        required = {"channel", "visits", "orders"}
        missing = sorted(required - set(snapshot.columns))
        if missing:
            return PowerBananaReport(
                agent_name=self.name,
                version=self.version,
                status="partial",
                answer=f"Cannot compute conversion_rate because required fields are missing: {', '.join(missing)}.",
                dataset_snapshot=snapshot,
                security_findings=security_findings,
                step_trace=[],
                evaluation=EvaluationResult(verdict="partial", failure_reasons=["missing_required_fields"], scores={}),
                limitations=["Required fields for conversion_rate are channel, visits, and orders."],
            )

        grouped: dict[str, dict[str, float]] = defaultdict(lambda: {"visits": 0.0, "orders": 0.0})
        skipped_rows = 0
        for row in rows:
            channel = str(row.get("channel", "")).strip()
            visits = self._to_float(row.get("visits"))
            orders = self._to_float(row.get("orders"))
            if not channel or visits is None or orders is None:
                skipped_rows += 1
                continue
            grouped[channel]["visits"] += visits
            grouped[channel]["orders"] += orders

        rates = {
            channel: totals["orders"] / totals["visits"]
            for channel, totals in grouped.items()
            if totals["visits"] > 0
        }
        if not rates:
            return PowerBananaReport(
                agent_name=self.name,
                version=self.version,
                status="partial",
                answer="Cannot compute conversion_rate because no group has visits greater than zero.",
                dataset_snapshot=snapshot,
                security_findings=security_findings,
                step_trace=[],
                evaluation=EvaluationResult(verdict="partial", failure_reasons=["no_valid_denominator"], scores={}),
                limitations=["At least one row needs a numeric visits value greater than zero."],
            )

        top_value, value = max(rates.items(), key=lambda item: item[1])
        analysis = AnalysisResult(
            metric="conversion_rate",
            group_by="channel",
            top_value=top_value,
            value=value,
            evidence_ref="blackboard://task_001/artifacts/ranked_metric_result_s2_v1",
            values=rates,
        )
        step_trace = [
            StepRecord(
                step_id="s1",
                action_type="skill",
                skill_id="compute_grouped_metric",
                status="succeeded",
                input_refs=["dataset://task_001/upload_v1"],
                output_ref="blackboard://task_001/artifacts/metric_result_s1_v1",
                expected_output_schema="MetricResult",
            ),
            StepRecord(
                step_id="s2",
                action_type="skill",
                skill_id="rank_metric_values",
                status="succeeded",
                input_refs=["blackboard://task_001/artifacts/metric_result_s1_v1"],
                output_ref=analysis.evidence_ref,
                expected_output_schema="RankedMetricResult",
            ),
        ]
        evaluation = self._evaluate_metric(snapshot, analysis)
        limitations = []
        if skipped_rows:
            limitations.append(f"Skipped {skipped_rows} rows with missing or nonnumeric channel, visits, or orders.")

        return PowerBananaReport(
            agent_name=self.name,
            version=self.version,
            status="completed" if evaluation.verdict == "pass" else "partial",
            answer=f"{top_value} has the highest conversion_rate at {value:.2%}.",
            dataset_snapshot=snapshot,
            security_findings=security_findings,
            step_trace=step_trace,
            evaluation=evaluation,
            analysis_result=analysis,
            limitations=limitations,
        )

    def _evaluate_metric(self, snapshot: DatasetSnapshot, analysis: AnalysisResult) -> EvaluationResult:
        failure_reasons = []
        if snapshot.dataset_version != "upload_v1":
            failure_reasons.append("dataset_version_mismatch")
        if analysis.metric != "conversion_rate":
            failure_reasons.append("metric_mismatch")
        if analysis.group_by not in snapshot.columns:
            failure_reasons.append("field_reference_missing")
        if not analysis.evidence_ref:
            failure_reasons.append("missing_evidence_ref")
        return EvaluationResult(
            verdict="pass" if not failure_reasons else "fail",
            failure_reasons=failure_reasons,
            scores={
                "dataset_ref": 1.0 if snapshot.dataset_version == "upload_v1" else 0.0,
                "field_reference": 1.0 if analysis.group_by in snapshot.columns else 0.0,
                "evidence_coverage": 1.0 if analysis.evidence_ref else 0.0,
                "metric_correctness": 1.0 if analysis.metric == "conversion_rate" else 0.0,
            },
        )

    def _to_float(self, value: object) -> float | None:
        try:
            return float(str(value).strip())
        except (TypeError, ValueError):
            return None

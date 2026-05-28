from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass
from pathlib import Path

from .blackboard import TaskBlackboard
from .models import AnalysisResult, DatasetSnapshot, EvaluationResult, PowerBananaReport, SecurityFinding
from .skills import compute_grouped_conversion_rate, conversion_rate_step_trace, evaluate_metric, rank_metric_values


PROMPT_INJECTION_PATTERNS = (
    re.compile(r"\bignore (all )?(previous|prior|above) instructions\b", re.IGNORECASE),
    re.compile(r"\breveal secrets?\b", re.IGNORECASE),
    re.compile(r"\bdisregard (the )?(system|developer) (message|instructions)\b", re.IGNORECASE),
)


@dataclass(frozen=True)
class SubAgentProfile:
    agent_id: str
    runtime_mode: str
    role_id: str
    autonomy_level: int | None = None


class DataProfileAgent:
    profile = SubAgentProfile("data_profile_agent", "workflow", "data_profiler")

    def run(self, blackboard: TaskBlackboard, path: Path) -> None:
        rows = self._load_rows(path)
        blackboard.rows = rows
        blackboard.dataset_snapshot = self._snapshot(path, rows)
        blackboard.security_findings = self._scan_security(rows)
        blackboard.artifacts["data_profile_v1"] = blackboard.dataset_snapshot
        blackboard.record_agent(
            self.profile.agent_id,
            self.profile.runtime_mode,
            "succeeded",
            "blackboard://task_001/artifacts/data_profile_v1",
        )

    def _load_rows(self, path: Path) -> list[dict[str, str]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", newline="", encoding="utf-8-sig") as handle:
                return list(csv.DictReader(handle))
        if suffix == ".xlsx":
            return self._load_xlsx_rows(path)
        raise ValueError(f"Unsupported file type: {suffix}. PowerBanana v0.1 supports .csv and simple .xlsx files.")

    def _load_xlsx_rows(self, path: Path) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Reading .xlsx files requires openpyxl. Install it or use CSV for PowerBanana v0.1.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value) if value is not None else "" for value in rows[0]]
        records: list[dict[str, str]] = []
        for row in rows[1:]:
            records.append(
                {
                    header: "" if value is None else str(value)
                    for header, value in zip(headers, row, strict=False)
                    if header
                }
            )
        return records

    def _snapshot(self, path: Path, rows: list[dict[str, str]]) -> DatasetSnapshot:
        file_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        columns = list(rows[0].keys()) if rows else []
        missing_counts = {
            column: sum(1 for row in rows if not str(row.get(column, "")).strip())
            for column in columns
        }
        return DatasetSnapshot(
            dataset_id=path.stem,
            dataset_version="upload_v1",
            file_hash=f"sha256:{file_hash}",
            row_count=len(rows),
            columns=columns,
            missing_counts=missing_counts,
        )

    def _scan_security(self, rows: list[dict[str, str]]) -> list[SecurityFinding]:
        findings: list[SecurityFinding] = []
        for row_index, row in enumerate(rows, start=2):
            for column, value in row.items():
                text = str(value)
                if any(pattern.search(text) for pattern in PROMPT_INJECTION_PATTERNS):
                    findings.append(
                        SecurityFinding(
                            risk_type="prompt_injection_in_cell",
                            source_ref=f"row:{row_index}:column:{column}",
                            action="exclude_as_instruction_keep_as_data",
                            detail=text,
                        )
                    )
        return findings


class DataAnalysisAgent:
    profile = SubAgentProfile("data_analysis_agent", "autonomous", "data_analyst", autonomy_level=2)

    def run(self, blackboard: TaskBlackboard) -> None:
        question = blackboard.question
        if self._is_ambiguous_performance_question(question):
            blackboard.status = "needs_clarification"
            blackboard.answer = "Please specify the metric to optimize, such as conversion_rate, revenue, orders, or visits."
            blackboard.evaluation = EvaluationResult(verdict="needs_clarification", failure_reasons=["ambiguous_metric"], scores={})
            blackboard.record_agent(self.profile.agent_id, self.profile.runtime_mode, "needs_clarification", "blackboard://task_001/decisions/clarification_required")
            return

        if not self._is_conversion_rate_question(question):
            blackboard.status = "needs_clarification"
            blackboard.answer = "PowerBanana v0.1 supports conversion-rate questions for CSV datasets. Please ask for conversion rate by group."
            blackboard.evaluation = EvaluationResult(verdict="needs_clarification", failure_reasons=["unsupported_question"], scores={})
            blackboard.record_agent(self.profile.agent_id, self.profile.runtime_mode, "needs_clarification", "blackboard://task_001/decisions/unsupported_question")
            return

        self._answer_conversion_rate(blackboard)

    def _answer_conversion_rate(self, blackboard: TaskBlackboard) -> None:
        snapshot = self._require_snapshot(blackboard)
        required = {"channel", "visits", "orders"}
        missing = sorted(required - set(snapshot.columns))
        if missing:
            blackboard.status = "partial"
            blackboard.answer = f"Cannot compute conversion_rate because required fields are missing: {', '.join(missing)}."
            blackboard.evaluation = EvaluationResult(verdict="partial", failure_reasons=["missing_required_fields"], scores={})
            blackboard.limitations = ["Required fields for conversion_rate are channel, visits, and orders."]
            blackboard.record_agent(self.profile.agent_id, self.profile.runtime_mode, "partial", "blackboard://task_001/evaluations/missing_required_fields")
            return

        rates, skipped_rows = compute_grouped_conversion_rate(blackboard.rows)
        if not rates:
            blackboard.status = "partial"
            blackboard.answer = "Cannot compute conversion_rate because no group has visits greater than zero."
            blackboard.evaluation = EvaluationResult(verdict="partial", failure_reasons=["no_valid_denominator"], scores={})
            blackboard.limitations = ["At least one row needs a numeric visits value greater than zero."]
            blackboard.record_agent(self.profile.agent_id, self.profile.runtime_mode, "partial", "blackboard://task_001/evaluations/no_valid_denominator")
            return

        top_value, value = rank_metric_values(rates)
        analysis = AnalysisResult(
            metric="conversion_rate",
            group_by="channel",
            top_value=top_value,
            value=value,
            evidence_ref="blackboard://task_001/artifacts/ranked_metric_result_s2_v1",
            values=rates,
        )
        blackboard.analysis_result = analysis
        blackboard.step_trace = conversion_rate_step_trace(analysis)
        blackboard.evaluation = evaluate_metric(snapshot, analysis)
        blackboard.answer = f"{top_value} has the highest conversion_rate at {value:.2%}."
        blackboard.status = "completed" if blackboard.evaluation.verdict == "pass" else "partial"
        if skipped_rows:
            blackboard.limitations.append(f"Skipped {skipped_rows} rows with missing or nonnumeric channel, visits, or orders.")
        blackboard.artifacts["analysis_result_v1"] = analysis
        blackboard.record_agent(self.profile.agent_id, self.profile.runtime_mode, blackboard.status, "blackboard://task_001/artifacts/analysis_result_v1")

    def _require_snapshot(self, blackboard: TaskBlackboard) -> DatasetSnapshot:
        if blackboard.dataset_snapshot is None:
            raise ValueError("data_analysis_agent requires data_profile_agent to create a dataset snapshot first.")
        return blackboard.dataset_snapshot

    def _is_ambiguous_performance_question(self, question: str) -> bool:
        q = question.lower()
        return "best" in q and "conversion" not in q and "revenue" not in q and "orders" not in q and "visits" not in q

    def _is_conversion_rate_question(self, question: str) -> bool:
        q = question.lower()
        return "conversion" in q and "rate" in q


class ReportAgent:
    profile = SubAgentProfile("report_agent", "workflow", "report_writer")

    def run(self, blackboard: TaskBlackboard, agent_name: str, version: str) -> PowerBananaReport:
        if blackboard.dataset_snapshot is None or blackboard.evaluation is None:
            raise ValueError("report_agent requires dataset_snapshot and evaluation before generating a report.")
        if blackboard.status == "completed":
            self._final_consistency_check(blackboard)
        blackboard.record_agent(self.profile.agent_id, self.profile.runtime_mode, "succeeded", "blackboard://task_001/artifacts/final_report_v1")
        return PowerBananaReport(
            agent_name=agent_name,
            version=version,
            status=blackboard.status,
            answer=blackboard.answer,
            dataset_snapshot=blackboard.dataset_snapshot,
            security_findings=blackboard.security_findings,
            agent_trace=blackboard.agent_trace,
            step_trace=blackboard.step_trace,
            evaluation=blackboard.evaluation,
            analysis_result=blackboard.analysis_result,
            limitations=blackboard.limitations,
        )

    def _final_consistency_check(self, blackboard: TaskBlackboard) -> None:
        analysis = blackboard.analysis_result
        snapshot = blackboard.dataset_snapshot
        if analysis is None or snapshot is None:
            raise ValueError("Cannot complete final report without analysis result and dataset snapshot.")
        if analysis.group_by not in snapshot.columns:
            blackboard.status = "partial"
            blackboard.limitations.append("Final consistency check failed: group_by field is missing from dataset snapshot.")


def build_default_subagent_registry() -> dict[str, SubAgentProfile]:
    agents = [DataProfileAgent(), DataAnalysisAgent(), ReportAgent()]
    return {agent.profile.agent_id: agent.profile for agent in agents}

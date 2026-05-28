from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from pathlib import Path

from .models import DatasetSnapshot, ToolCallRecord


@dataclass(frozen=True)
class DatasetReadResult:
    rows: list[dict[str, str]]
    snapshot: DatasetSnapshot
    tool_call: ToolCallRecord


class ToolGateway:
    def read_dataset_snapshot(self, path: Path) -> DatasetReadResult:
        rows = self._load_rows(path)
        snapshot = self._snapshot(path, rows)
        output_ref = "dataset://task_001/upload_v1"
        return DatasetReadResult(
            rows=rows,
            snapshot=snapshot,
            tool_call=ToolCallRecord(
                tool_id="dataset.read_snapshot",
                status="succeeded",
                risk_level="low",
                input_ref=str(path),
                output_ref=output_ref,
            ),
        )

    def _load_rows(self, path: Path) -> list[dict[str, str]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", newline="", encoding="utf-8-sig") as handle:
                return list(csv.DictReader(handle))
        if suffix == ".xlsx":
            return self._load_xlsx_rows(path)
        raise ValueError(f"Unsupported file type: {suffix}. PowerBanana v0.1 supports .csv and simple .xlsx files.")

    def _load_xlsx_rows(self, path: Path) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Reading .xlsx files requires openpyxl. Install it or use CSV for PowerBanana v0.1.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value) if value is not None else "" for value in rows[0]]
        records: list[dict[str, str]] = []
        for row in rows[1:]:
            records.append(
                {
                    header: "" if value is None else str(value)
                    for header, value in zip(headers, row, strict=False)
                    if header
                }
            )
        return records

    def _snapshot(self, path: Path, rows: list[dict[str, str]]) -> DatasetSnapshot:
        file_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        columns = list(rows[0].keys()) if rows else []
        missing_counts = {
            column: sum(1 for row in rows if not str(row.get(column, "")).strip())
            for column in columns
        }
        return DatasetSnapshot(
            dataset_id=path.stem,
            dataset_version="upload_v1",
            file_hash=f"sha256:{file_hash}",
            row_count=len(rows),
            columns=columns,
            missing_counts=missing_counts,
        )
